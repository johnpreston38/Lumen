import glob
import json
import os
import sys
import time
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

# Для нового Telegram-бота укажи отдельный токен.
# Пример:
# DATALENS_BOT_TOKEN="123:ABC" DATALENS_CHAT_ID="105015582" python3 datalens_daily_bot.py --once
DATALENS_BOT_TOKEN = os.getenv("DATALENS_BOT_TOKEN", "")
DATALENS_CHAT_ID = os.getenv("DATALENS_CHAT_ID", "")

SEND_TIME = os.getenv("DATALENS_SEND_TIME", "10:05")
CHECK_INTERVAL_SECONDS = int(os.getenv("DATALENS_CHECK_INTERVAL_SECONDS", "900"))
STATE_DIR = Path(os.getenv(
    "DATALENS_STATE_DIR",
    "/Users/evgenijnovickij/Library/Application Support/LumenBots",
))
STATE_DIR.mkdir(parents=True, exist_ok=True)
LAST_SENT_FILE = STATE_DIR / "datalens_daily_bot.last_sent_day"
UPDATE_OFFSET_FILE = STATE_DIR / "datalens_daily_bot.telegram_offset"

EXPORT_XLSX = os.getenv("DATALENS_EXPORT_XLSX", "")
DEFAULT_EXPORT_PATTERN = "/Users/evgenijnovickij/Downloads/Табл*.xlsx"

REQUIRED_COLUMNS = {
    "period": "DateTrunc1",
    "percap_total": "Перкап Итого",
    "foodcost": "Фудкост, %",
    "viewers": "Зрители",
    "work_hours": "ЧасыРаботы",
    "total_revenue": "Итого",
    "payroll": "Начислено",
    "lc": "LC",
    "bar_share": "ДоляБара",
}


def latest_export_file():
    if EXPORT_XLSX:
        return EXPORT_XLSX

    files = glob.glob(DEFAULT_EXPORT_PATTERN)

    if not files:
        raise FileNotFoundError(
            f"Не найден XLSX-файл выгрузки по маске: {DEFAULT_EXPORT_PATTERN}"
        )

    return max(files, key=os.path.getmtime)


def parse_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
    cleaned = "".join(ch for ch in cleaned if ch.isdigit() or ch in ".-")

    if not cleaned:
        return None

    return float(cleaned)


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "")


def read_latest_metrics(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("В XLSX нет строк с данными.")

    headers = [str(value or "").strip() for value in rows[0]]
    header_map = {
        normalize_header(header): index
        for index, header in enumerate(headers)
    }

    column_indexes = {}

    for key, column_name in REQUIRED_COLUMNS.items():
        index = header_map.get(normalize_header(column_name))

        if index is None:
            raise ValueError(f"Не найдена колонка: {column_name}")

        column_indexes[key] = index

    data_rows = [row for row in rows[1:] if any(value is not None for value in row)]

    if not data_rows:
        raise ValueError("В XLSX нет строк с показателями.")

    def row_period(row):
        value = row[column_indexes["period"]]

        if isinstance(value, datetime):
            return value

        try:
            return datetime.fromisoformat(str(value).replace("Z", ""))
        except ValueError:
            return datetime.min

    latest_row = max(data_rows, key=row_period)
    period = row_period(latest_row)
    month_rows = [
        row for row in data_rows
        if row_period(row).year == period.year and row_period(row).month == period.month
    ]

    def value_from_row(row, key):
        return parse_number(row[column_indexes[key]])

    def average_for_month(key):
        values = [value_from_row(row, key) for row in month_rows]
        values = [value for value in values if value is not None]

        if not values:
            return None

        return sum(values) / len(values)

    def average_for_month_day_type(key):
        current_is_weekend = period.weekday() >= 5
        values = []

        for row in month_rows:
            current_period = row_period(row)

            if (current_period.weekday() >= 5) != current_is_weekend:
                continue

            value = value_from_row(row, key)

            if value is not None:
                values.append(value)

        if not values:
            return None

        return sum(values) / len(values)

    return {
        "source_file": path,
        "period": period,
        "percap_total": value_from_row(latest_row, "percap_total"),
        "foodcost": value_from_row(latest_row, "foodcost"),
        "viewers": value_from_row(latest_row, "viewers"),
        "work_hours": value_from_row(latest_row, "work_hours"),
        "total_revenue": value_from_row(latest_row, "total_revenue"),
        "payroll": value_from_row(latest_row, "payroll"),
        "lc": value_from_row(latest_row, "lc"),
        "bar_share": value_from_row(latest_row, "bar_share"),
        "month_avg": {
            "percap_total": average_for_month("percap_total"),
            "foodcost": average_for_month("foodcost"),
            "viewers": average_for_month("viewers"),
            "viewers_same_day_type": average_for_month_day_type("viewers"),
            "work_hours": average_for_month("work_hours"),
            "total_revenue": average_for_month("total_revenue"),
            "payroll": average_for_month("payroll"),
            "lc": average_for_month("lc"),
            "bar_share": average_for_month("bar_share"),
        },
    }


def collect_metrics():
    return read_latest_metrics(latest_export_file())


def previous_month_daily_average(current_path, period, metric_key):
    previous_month = period.month - 1
    previous_year = period.year

    if previous_month == 0:
        previous_month = 12
        previous_year -= 1

    candidates = sorted(
        Path("/Users/evgenijnovickij/Downloads").glob("Табл*.xlsx"),
        key=lambda file_path: file_path.stat().st_mtime,
        reverse=True,
    )

    for candidate in candidates:
        try:
            workbook = load_workbook(candidate, data_only=True, read_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [str(value or "").strip() for value in rows[0]]
            header_map = {
                normalize_header(header): index
                for index, header in enumerate(headers)
            }
            period_index = header_map.get(normalize_header(REQUIRED_COLUMNS["period"]))
            metric_index = header_map.get(normalize_header(REQUIRED_COLUMNS[metric_key]))

            if period_index is None or metric_index is None:
                continue

            for row in rows[1:]:
                row_period_value = row[period_index]

                if not isinstance(row_period_value, datetime):
                    continue

                if row_period_value.year != previous_year or row_period_value.month != previous_month:
                    continue

                value = parse_number(row[metric_index])

                if value is None:
                    continue

                # Если строка месячная, делим часы на количество дней месяца,
                # чтобы получить сопоставимый дневной ориентир.
                days_in_previous_month = (
                    datetime(period.year, period.month, 1) - timedelta(days=1)
                ).day
                return value / days_in_previous_month
        except Exception:
            continue

    return None


def rub(value):
    if value is None:
        return "н/д"

    return f"{value:,.0f} ₽".replace(",", " ")


def hours(value):
    if value is None:
        return "н/д"

    return f"{value:,.1f} ч".replace(",", " ")


def percent(value):
    if value is None:
        return "н/д"

    if abs(value) <= 1:
        value *= 100

    return f"{value:.1f}%"


def delta_percent(current, average, lower_is_better=False):
    if current is None or average in (None, 0):
        return "⚪ н/д"

    delta = (current - average) / average * 100
    is_good = delta >= 0

    if lower_is_better:
        is_good = delta <= 0

    emoji = "🟢" if is_good else "🔴"
    sign = "+" if delta > 0 else ""
    return f"{emoji} {sign}{delta:.1f}%"


def count_value(value):
    if value is None:
        return "н/д"

    return f"{value:,.0f}".replace(",", " ")


def build_report():
    metrics = collect_metrics()
    today = datetime.now().strftime("%d.%m.%Y")
    period = metrics["period"].strftime("%d.%m.%Y")
    month_avg = metrics["month_avg"]
    viewers_label = "среднее за месяц по выходным" if metrics["period"].weekday() >= 5 else "среднее за месяц по будням"

    message = f"📊 ЛЮМЕН — ЕЖЕДНЕВНЫЙ СРЕЗ\n{today}\n"
    message += f"Период в выгрузке: {period}\n\n"

    data_lag_days = (datetime.now().date() - metrics["period"].date()).days

    if data_lag_days > 2:
        message += f"⚠️ Внимание: последняя найденная выгрузка старше {data_lag_days} дней.\n\n"

    message += "Ключевые показатели\n"
    message += (
        f"• Перкап итого: {rub(metrics['percap_total'])} | "
        f"среднее за месяц: {rub(month_avg['percap_total'])} | "
        f"{delta_percent(metrics['percap_total'], month_avg['percap_total'])}\n"
    )
    message += (
        f"• Фудкост: {percent(metrics['foodcost'])} | "
        f"среднее за месяц: {percent(month_avg['foodcost'])} | "
        f"{delta_percent(metrics['foodcost'], month_avg['foodcost'], lower_is_better=True)}\n"
    )
    message += (
        f"• Зрители: {count_value(metrics['viewers'])} | "
        f"{viewers_label}: {count_value(month_avg['viewers_same_day_type'])} | "
        f"{delta_percent(metrics['viewers'], month_avg['viewers_same_day_type'])}\n"
    )
    message += (
        f"• Часы работы: {hours(metrics['work_hours'])}\n"
    )
    message += (
        f"• LC: {percent(metrics['lc'])} | "
        f"среднее за месяц: {percent(month_avg['lc'])} | "
        f"{delta_percent(metrics['lc'], month_avg['lc'], lower_is_better=True)}\n"
    )
    message += (
        f"• Доля бара: {percent(metrics['bar_share'])} | "
        f"среднее за месяц: {percent(month_avg['bar_share'])} | "
        f"{delta_percent(metrics['bar_share'], month_avg['bar_share'])}\n"
    )

    return message


def resolve_chat_id():
    if DATALENS_CHAT_ID:
        return DATALENS_CHAT_ID

    if not DATALENS_BOT_TOKEN:
        raise ValueError("Не указан DATALENS_BOT_TOKEN для нового Telegram-бота.")

    url = f"https://api.telegram.org/bot{DATALENS_BOT_TOKEN}/getUpdates"

    with urllib.request.urlopen(url, timeout=30) as response:
        data = json.loads(response.read().decode("utf-8"))

    for update in reversed(data.get("result", [])):
        message = update.get("message") or update.get("channel_post") or {}
        chat = message.get("chat") or {}
        chat_id = chat.get("id")

        if chat_id:
            return str(chat_id)

    raise ValueError(
        "Не удалось определить chat_id. Напиши новому боту /start и запусти снова "
        "или укажи DATALENS_CHAT_ID вручную."
    )


def telegram_request(method, payload=None):
    if not DATALENS_BOT_TOKEN:
        raise ValueError("Не указан DATALENS_BOT_TOKEN для нового Telegram-бота.")

    url = f"https://api.telegram.org/bot{DATALENS_BOT_TOKEN}/{method}"

    if payload is None:
        with urllib.request.urlopen(url, timeout=35) as response:
            result = json.loads(response.read().decode("utf-8"))
    else:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(
            url,
            data=data,
            method="POST",
            headers={"Content-Type": "application/json; charset=utf-8"},
        )

        with urllib.request.urlopen(request, timeout=35) as response:
            result = json.loads(response.read().decode("utf-8"))

    if not result.get("ok"):
        raise RuntimeError(f"Telegram вернул ошибку: {result}")

    return result


def report_keyboard():
    return {
        "inline_keyboard": [[
            {"text": "📊 Получить срез", "callback_data": "run_report"}
        ]]
    }


def send_telegram_message(text, chat_id=None, with_button=False):
    chat_id = chat_id or resolve_chat_id()
    payload = {
        "chat_id": chat_id,
        "text": text,
        "disable_web_page_preview": True,
    }

    if with_button:
        payload["reply_markup"] = report_keyboard()

    telegram_request("sendMessage", payload)

    print(f"Отчет отправлен в chat_id={chat_id}", flush=True)


def send_start_message(chat_id):
    send_telegram_message(
        "Готов. Нажми кнопку, чтобы получить свежий срез по последней XLSX-выгрузке.",
        chat_id=chat_id,
        with_button=True,
    )


def answer_callback(callback_query_id, text="Запускаю отчет"):
    try:
        telegram_request("answerCallbackQuery", {
            "callback_query_id": callback_query_id,
            "text": text,
        })
    except Exception as error:
        print(f"Не удалось ответить на callback: {error}", flush=True)


def read_update_offset():
    try:
        return int(UPDATE_OFFSET_FILE.read_text(encoding="utf-8").strip())
    except (FileNotFoundError, ValueError):
        return 0


def write_update_offset(offset):
    UPDATE_OFFSET_FILE.write_text(str(offset), encoding="utf-8")


def handle_update(update):
    message = update.get("message") or {}
    callback_query = update.get("callback_query") or {}

    if message:
        chat = message.get("chat") or {}
        chat_id = chat.get("id")
        text = (message.get("text") or "").strip()

        if chat_id and text in ("/start", "/menu", "start"):
            send_start_message(chat_id)
        elif chat_id and text in ("/report", "📊 Получить срез"):
            send_telegram_message(build_report(), chat_id=chat_id, with_button=True)

    if callback_query:
        callback_id = callback_query.get("id")
        data = callback_query.get("data")
        message = callback_query.get("message") or {}
        chat = message.get("chat") or {}
        chat_id = chat.get("id")

        if callback_id:
            answer_callback(callback_id)

        if chat_id and data == "run_report":
            send_telegram_message(build_report(), chat_id=chat_id, with_button=True)


def run_bot_listener():
    print("Telegram button listener started.", flush=True)

    while True:
        try:
            offset = read_update_offset()
            result = telegram_request("getUpdates", {
                "offset": offset,
                "timeout": 25,
                "allowed_updates": ["message", "callback_query"],
            })

            for update in result.get("result", []):
                update_id = update.get("update_id")

                if update_id is not None:
                    write_update_offset(update_id + 1)

                handle_update(update)
        except Exception as error:
            print(f"Ошибка Telegram listener: {error}", flush=True)
            time.sleep(10)


def run_once(send=True):
    message = build_report()
    print(message)

    if send:
        send_telegram_message(message)


def already_sent_today():
    today = datetime.now().strftime("%Y-%m-%d")

    try:
        return LAST_SENT_FILE.read_text(encoding="utf-8").strip() == today
    except FileNotFoundError:
        return False


def mark_sent_today():
    LAST_SENT_FILE.write_text(datetime.now().strftime("%Y-%m-%d"), encoding="utf-8")


def next_send_datetime():
    hour, minute = map(int, SEND_TIME.split(":"))
    now = datetime.now()
    next_send = now.replace(hour=hour, minute=minute, second=0, microsecond=0)

    if next_send <= now:
        next_send += timedelta(days=1)

    return next_send


def should_send_now():
    hour, minute = map(int, SEND_TIME.split(":"))
    now = datetime.now()
    planned = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    return now >= planned and not already_sent_today()


def run_daily():
    while True:
        if not should_send_now():
            wait_seconds = min(
                CHECK_INTERVAL_SECONDS,
                max(1, int((next_send_datetime() - datetime.now()).total_seconds())),
            )
            print(f"Следующая проверка расписания: через {wait_seconds // 60} мин.", flush=True)
            time.sleep(wait_seconds)
            continue

        try:
            run_once(send=True)
            mark_sent_today()
        except Exception as error:
            print(f"Ошибка при отправке отчета: {error}", flush=True)
            time.sleep(60 * 10)


if __name__ == "__main__":
    if "--once" in sys.argv:
        run_once(send=True)
    elif "--preview" in sys.argv:
        run_once(send=False)
    elif "--bot" in sys.argv:
        run_bot_listener()
    else:
        run_daily()
